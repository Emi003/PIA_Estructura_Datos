import sqlite3
from sqlite3 import Error
import csv
import datetime
import pandas as pd
import sys
from openpyxl import Workbook
from openpyxl.styles import Alignment

MenuPrincipal = None

def validar_rfc_persona_fisica(rfc):
    rfc = rfc.strip().upper()
    if len(rfc) != 13:
        return False
    if not rfc[:4].isalpha():
        return False
    if not rfc[4:10].isdigit():
        return False
    if not (rfc[-1].isdigit() or rfc[-1] == 'A'):
        return False
    homoclave = rfc[10:12]
    if not homoclave.isalnum():
        return False
    rfc_sin_homoclave = rfc[:10] + rfc[11]
    return True


def validar_rfc_persona_moral(rfc):
    rfc = rfc.strip().upper()
    if len(rfc) != 12:
        return False
    if not rfc[:3].isalpha():
        return False
    if not rfc[3:9].isdigit():
        return False
    if not (rfc[-1].isdigit() or rfc[-1] == 'A'):
        return False
    homoclave = rfc[9:11]
    if not homoclave.isalnum():
        return False
    rfc_sin_homoclave = rfc[:9] + rfc[11]
    return True


def validar_correo(correo):
    if correo.count('@') != 1:
        return False
    usuario, dominio = correo.split('@')
    if not usuario or not dominio:
        return False
    if dominio.count('.') != 1 or not dominio.split('.')[1]:
        return False
    return True




while MenuPrincipal != "4":
    print("Menú Principal")
    print("1. Notas")
    print("2. Clientes")
    print("3. Servicios")
    print("4. Estadistica
    print("5. Salir")

    MenuPrincipal = input("Seleccione una opción: ")
    if MenuPrincipal == "1":
        try:
            conn = sqlite3.connect("Evidencia3.db")
            mi_cursor = conn.cursor()

            while True:
                print("Menú Notas:")
                print("1. Registrar una nota")
                print("2. Cancelar una nota")
                print("3. Recuperar una nota")
                print("4. Consultas y reportes de notas")
                print("5. Volver al menú principal")
                
                opcion = input("Ingrese la opción deseada: ")

                if opcion == "1":
                    mi_cursor.execute("SELECT ClaveC, NombreCompleto FROM Clientes")
                    clientes = mi_cursor.fetchall()
                    print("Clientes registrados:")
                    for ClaveC, NombreCompleto in clientes:
                        print(f"Clave: {ClaveC}, Nombre: {NombreCompleto}")

                    clave_cliente = input("Ingrese la clave del cliente al cual se expedirá la nota: ")

                    fecha_nota = input("Dame la fecha en este formato: dd/mm/aaaa: ")
                    try:
                        fecha_actual = datetime.date.today()
                        fecha_procesada = datetime.datetime.strptime(fecha_nota, '%d/%m/%Y').date()
                        if fecha_procesada > fecha_actual:
                            print("La fecha no puede ser posterior a la fecha actual.")
                            continue
                    except ValueError:
                        print("Formato de fecha incorrecto. Introduce la fecha en el formato dd/mm/aaaa.")
                        continue

                    detalles_nota = []

                    while True:
                        mi_cursor.execute("SELECT ClaveS, Nombre, Costo FROM Servicios")
                        servicios = mi_cursor.fetchall()
                        print("Servicios registrados:")
                        for ClaveS, Nombre, Costo in servicios:
                            print(f"Clave: {ClaveS}, Nombre: {Nombre}, Costo: {Costo}")

                        clave_servicio = input("Ingrese la clave del servicio a agregar a la nota: ")

                        servicio_seleccionado = next((servicio for servicio in servicios if str(servicio[0]) == clave_servicio), None)

                        if servicio_seleccionado:
                            detalles_nota.append(servicio_seleccionado)
                        else:
                            print("Clave de servicio no válida. Por favor, ingrese una clave válida.")

                        agregar_otro = input("¿Desea agregar otro servicio a la nota? (SI/NO): ")
                        if agregar_otro.upper() != "SI":
                            break


                    monto_total = sum(servicio[2] for servicio in detalles_nota)

                    datos_nota = (clave_cliente, fecha_procesada)
                    mi_cursor.execute("INSERT INTO Notas (ClaveC, Fecha) VALUES (?, ?)", datos_nota)
                    conn.commit()

                    mi_cursor.execute("SELECT last_insert_rowid()")
                    last_id = mi_cursor.fetchone()[0]

                    for clave_servicio, nombre, Costo in detalles_nota: 
                        datos_detalle = (last_id, clave_servicio, Costo)
                        mi_cursor.execute("INSERT INTO DetalleNotas (Folio, ClaveS, Monto) VALUES (?, ?, ?)", datos_detalle)
                        conn.commit()

                    print("Registro de la nota agregado exitosamente")



                elif opcion == "2":
                    folio_cancelar = input("Ingrese el folio de la nota que desea cancelar: ")
                    
                    mi_cursor.execute("SELECT Notas.Folio, ClaveC, Fecha, ClaveS, Monto FROM Notas JOIN DetalleNotas ON Notas.Folio = DetalleNotas.Folio WHERE Notas.Folio = ? AND Estado_Nota = 0", (folio_cancelar,))
                    nota = mi_cursor.fetchall()  
                    
                    if nota:
                        print("Datos de la nota a cancelar:")
                        print(f"Folio: {nota[0][0]}")
                        print(f"Clave del cliente: {nota[0][1]}")
                        print(f"Fecha: {nota[0][2]}")
                        

                        monto_total = 0
                        

                        for servicio in nota:
                            clave_servicio, monto = servicio[3], servicio[4]
                            print(f"Clave del servicio: {clave_servicio}, Monto: {monto}")
                            monto_total += monto
                        
                        print(f"Monto Total: {monto_total}")
                        
                        confirmar_cancelacion = input("¿Desea cancelar esta nota? (SI/NO): ")
                        
                        if confirmar_cancelacion.upper() == "SI":
                            # Marcar la nota como cancelada
                            mi_cursor.execute("UPDATE Notas SET Estado_Nota = 1 WHERE Folio = ?", (folio_cancelar,))
                            conn.commit()
                            print("La nota ha sido cancelada exitosamente.")
                        else:
                            print("La nota no ha sido cancelada.")
                    else:
                        print("La nota no existe o ya ha sido cancelada.")


                elif opcion == "3":

                    mi_cursor.execute("SELECT Folio, ClaveC, Fecha FROM Notas WHERE Estado_Nota = 1")
                    notas_canceladas = mi_cursor.fetchall()

                    if not notas_canceladas:
                        print("No hay notas canceladas para recuperar.")
                    else:
                        print("Listado de notas canceladas:")
                        print("Folio\tClave del Cliente\tFecha")
                        for nota in notas_canceladas:
                            print(f"{nota[0]}\t{nota[1]}\t{nota[2]}")

                        folio_recuperar = input("Ingrese el folio de la nota que desea recuperar (o 'NO' para cancelar): ")

                        if folio_recuperar.upper() != 'NO':
                            mi_cursor.execute("SELECT Folio, ClaveC, Fecha FROM Notas WHERE Folio = ? AND Estado_Nota = 1", (folio_recuperar,))
                            nota = mi_cursor.fetchone()

                            if nota:
                                print("Datos de la nota a recuperar:")
                                print(f"Folio: {nota[0]}")
                                print(f"Clave del Cliente: {nota[1]}")
                                print(f"Fecha: {nota[2]}")

                                confirmar_recuperacion = input("¿Desea recuperar esta nota? (SI/NO): ")

                                if confirmar_recuperacion.upper() == "SI":
                                    mi_cursor.execute("UPDATE Notas SET Estado_Nota = 0 WHERE Folio = ?", (folio_recuperar,))
                                    conn.commit()
                                    print("La nota ha sido recuperada exitosamente.")
                                else:
                                    print("La nota no ha sido recuperada.")
                            else:
                                print("La nota no existe o no está cancelada.")


                elif opcion == "4":
                    while True:
                        print("Consultas y reportes de notas:")
                        print("1. Consulta por período")
                        print("2. Consulta por folio")
                        print("3. Volver al menú Notas")
                        
                        sub_opcion = input("Ingrese la opción deseada: ")

                        if sub_opcion == "1":
                            try:
                                with sqlite3.connect("Evidencia3.db",
                                                    detect_types = sqlite3.PARSE_DECLTYPES | sqlite3.PARSE_COLNAMES) as conn:
                                    fecha_inicial = input("Ingrese la fecha inicial en formato dd/mm/aaaa (o presione Enter para utilizar 01/01/2000): ")
                                    fecha_final = input("Ingrese la fecha final en formato dd/mm/aaaa (o presione Enter para utilizar la fecha actual): ")

                                    try:
                                        if not fecha_inicial:
                                            fecha_inicial = "01/01/2000"
                                        if not fecha_final:
                                            fecha_final = datetime.datetime.now().strftime("%d/%m/%Y")

                                        fecha_inicial = datetime.datetime.strptime(fecha_inicial, '%d/%m/%Y').date()
                                        fecha_final = datetime.datetime.strptime(fecha_final, '%d/%m/%Y').date()

                                        if fecha_inicial > fecha_final:
                                            print("La fecha inicial no puede ser posterior a la fecha final.")
                                        else:
                                            mi_cursor.execute("SELECT Notas.Folio, strftime('%d/%m/%Y', Notas.Fecha), Clientes.NombreCompleto FROM Notas JOIN Clientes ON Notas.ClaveC = Clientes.ClaveC WHERE Notas.Fecha BETWEEN ? AND ? AND Estado_Nota = 0", (fecha_inicial, fecha_final))

                                            notas_periodo = mi_cursor.fetchall()

                                            if not notas_periodo:
                                                print("No hay notas emitidas para el período seleccionado.")
                                            else:
                                                print("Notas emitidas en el período seleccionado:")
                                                for nota in notas_periodo:
                                                    print(f"Folio: {nota[0]}, Fecha: {nota[1]}, Cliente: {nota[2]}")

                                                monto_promedio = sum(nota[0] for nota in notas_periodo) / len(notas_periodo) if len(notas_periodo) > 0 else 0

                                                print(f"Monto promedio de las notas en el período: {monto_promedio}")
                                                exportar_reporte = input("¿Desea exportar este resultado a CSV o Excel? (CSV/Excel/NO): ")
                                                if exportar_reporte.upper() == "CSV":
                                                    nombre_archivo = "ReportePorPeriodo_{}_{}.csv".format(fecha_inicial.strftime("%d_%m_%Y"), fecha_final.strftime("%m_%d_%Y"))
                                                    with open(nombre_archivo, 'w', newline='') as archivo_csv:
                                                        campo_nombres = ['Folio', 'Fecha', 'Cliente']
                                                        escritor_csv = csv.DictWriter(archivo_csv, fieldnames=campo_nombres)

                                                        escritor_csv.writeheader()
                                                        for nota in notas_periodo:
                                                            escritor_csv.writerow({'Folio': nota[0], 'Fecha': nota[1], 'Cliente': nota[2]})

                                                    print(f"Los datos se han exportado a '{nombre_archivo}' exitosamente.")
                                                elif exportar_reporte.upper() == "EXCEL":
                                                    nombre_archivo = "ReportePorPeriodo_{}_{}.xlsx".format(fecha_inicial.strftime("%d_%m_%Y"), fecha_final.strftime("%m_%d_%Y"))
                                                    workbook = Workbook()
                                                    sheet = workbook.active

                                                    sheet.append(['Folio', 'Fecha', 'Cliente'])
                                                    for nota in notas_periodo:
                                                        sheet.append([nota[0], nota[1], nota[2]])

                                                    workbook.save(nombre_archivo)
                                                    print(f"Los datos se han exportado a '{nombre_archivo}' exitosamente.")
                                                elif exportar_reporte.upper() == "NO":
                                                    continue
                                                else:
                                                    print("Opción de exportación no válida.")
                                    except ValueError:
                                        print("Formato de fecha incorrecto. Introduce la fecha en el formato dd/mm/aaaa.")

                            finally:
                                conn.close

                        elif sub_opcion == "2":
                            mi_cursor.execute("SELECT Notas.Folio, Notas.Fecha, Clientes.NombreCompleto FROM Notas JOIN Clientes ON Notas.ClaveC = Clientes.ClaveC WHERE Estado_Nota = 0 ORDER BY Notas.Folio")
                            notas = mi_cursor.fetchall()

                            print("Notas registradas:")
                            for nota in notas:
                                print(f"Folio: {nota[0]}, Fecha: {nota[1]}, Cliente: {nota[2]}")

                            folio_consultar = input("Ingrese el folio de la nota que desea consultar: ")
                            nota_seleccionada = next((nota for nota in notas if str(nota[0]) == folio_consultar), None)

                            if nota_seleccionada:
                                folio, fecha, nombre_cliente = nota_seleccionada
                                print("Detalle de la nota seleccionada:")
                                print(f"Folio: {folio}")
                                print(f"Fecha: {fecha}")
                                print(f"Cliente: {nombre_cliente}")

                                mi_cursor.execute("SELECT DetalleNotas.ClaveS, Servicios.Nombre, DetalleNotas.Monto FROM DetalleNotas JOIN Servicios ON DetalleNotas.ClaveS = Servicios.ClaveS WHERE Folio = ?", (folio,))
                                detalle_nota = mi_cursor.fetchall()

                                if not detalle_nota:
                                    print("Esta nota no tiene servicios asociados.")
                                else:
                                    print("Detalle de servicios:")
                                    for servicio in detalle_nota:
                                        clave_servicio, nombre_servicio, monto_servicio = servicio
                                        print(f"Clave del servicio: {clave_servicio}, Nombre del servicio: {nombre_servicio}, Monto: {monto_servicio}")
                            else:
                                print("El folio ingresado no corresponde a una nota válida.")

                        elif sub_opcion == "3":
                            break
                        else:
                            print("Opción no válida. Por favor, seleccione una opción válida.")

                elif opcion == "5":
                    break
                else:
                    print("Opción no válida. Por favor, seleccione una opción válida.")
            
        except Error as e:
            print(e)
        except Exception:
            print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
        finally:
            conn.close()


    elif MenuPrincipal == "2":
        while True:
            print("Menú Clientes:")
            print("1. Agregar un Cliente")
            print("2. Cancelar un Cliente")
            print("3. Recuperar un Cliente")
            print("4. Consultas y reportes de clientes")
            print("3. Volver al menú principal")
            
            opcion = input("Ingrese la opción deseada: ")

            if opcion == "1":
                try:
                    with sqlite3.connect("Evidencia3.db") as conn:
                        mi_cursor = conn.cursor()
                        nombre = input("Ingrese el nombre completo del cliente: ")
                        if not nombre.strip():
                            print("El nombre no puede quedar vacío ni ser solo espacios en blanco.")
                            continue
                        tipo_cliente = input("¿Es persona física o moral? (F/M): ").strip().upper() 
                        rfc = input("Introduce el RFC del cliente: ")
                        if tipo_cliente == "F":
                            if not validar_rfc_persona_fisica(rfc):
                                print("RFC de persona física no válido.")
                                continue
                        elif tipo_cliente == "M":
                            if not validar_rfc_persona_moral(rfc):
                                print("RFC de persona moral no válido.")
                                continue
                        else:
                            print("Opción no válida. Debe seleccionar F para persona física o M para persona moral.")
                            continue

                        correo = input("Ingrese el correo electrónico del cliente: ")
                        if not validar_correo(correo):
                            print("Correo electrónico no válido.")
                            continue


                        mi_cursor.execute("INSERT INTO Clientes (NombreCompleto, RFC, CORREO) VALUES (?, ?, ?)",
                                    (nombre, rfc, correo))
                        conn.commit()
                        print("El cliente ha sido registrado con éxito.")
                except Error as e:
                    print(e)
                except Exception:
                    print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                finally:
                    conn.close


            elif opcion == "2":
              mi_cursor.execute("SELECT ClaveC, NombreCompleto FROM Clientes WHERE Estado_Cliente = 0", (cliente_cancelar,))
              cliente = mi_cursor.fetchall()
              if not cliente:
                print("No hay clientes registrados en este momento.")
              else:
                print("Clientes Suspendido: ")
                for ClaveC, NombreCompleto in cliente:
                  print(f"Clave: {ClaveC}, Nombre: {NombreCompleto}")
                cliente_cancelar = input("Ingrese la clave del cliente desea Suspender: ")
                
                if cliente_cancelar != "0":
                  mi_cursor.execute("SELECT * FROM Clientes WHERE ClaveC = ? AND Estado_Cliente = 1", (cliente_cancelar,))
                  cliente = mi_cursor.fetchone()
                  if cliente:
                    print("Datos del cliente a suspender: ")
                    print(f"Clave: {cliente_suspender[0]}")
                    print(f"Nombre Completo: {cliente_suspender[1]}")
                    print(f"RFC: {cliente_suspender[2]}")
                    print(f"Correo Electrónico: {cliente_suspender[3]}")
                  
                  
                    confirmar_cancelacion = input("¿Desea cancelar esta nota? (SI/NO): ")
                    
                    if confirmar_cancelacion.upper() == "SI":
                        mi_cursor.execute("UPDATE Clientes SET Estado_Cliente = 1 WHERE ClaveC = ?", (folio_cancelar,))
                        conn.commit()
                        print("La nota ha sido cancelada exitosamente.")
                    else:
                        print("La nota no ha sido cancelada.")
                else:
                    print("La nota no existe o ya ha sido cancelada.")
                      

            elif opcion == "3":
              mi_cursor.execute("SELECT ClaveC, NombreCompleto FROM Clientes WHERE Estado_Cliente = 0")
              clientes_recuperar = mi_cursor.fetchall()
      
              if not clientes_recuperar:
                  print("No hay clientes suspendidos en este momento.")
              else:
                  print("Clientes suspendidos:")
                  for clave, nombre in clientes_recuperar:
                      print(f"Clave: {clave}, Nombre: {nombre}")
      
                  clave_recuperar = input("Ingrese la clave del cliente que desea recuperar (o 0 para volver al menú anterior): ")
      
                  if clave_recuperar != "0":
                      mi_cursor.execute("SELECT * FROM Clientes WHERE ClaveC = ? AND Estado_Cliente = 1", (clave_recuperar,))
                      cliente_recuperar = mi_cursor.fetchone()
      
                      if cliente_recuperar:
                          print("Datos del cliente a recuperar:")
                          print(f"Clave: {cliente_recuperar[0]}")
                          print(f"Nombre Completo: {cliente_recuperar[1]}")
                          print(f"RFC: {cliente_recuperar[2]}")
                          print(f"Correo Electrónico: {cliente_recuperar[3]}")
      
                          confirmacion = input("¿Está seguro de que desea recuperar a este cliente? (SI/NO): ")
                          if confirmacion.upper() == "SI":
                              mi_cursor.execute("UPDATE Clientes SET Estado_Cliente = 0 WHERE ClaveC = ?", (clave_recuperar,))
                              print("Cliente recuperado exitosamente. Ahora puede registrar nuevas notas.")
                          elif confirmacion.upper() == "NO":
                              print("Operación cancelada.")
                          else:
                              print("Opción no válida.")
                      else:
                          print("Cliente no encontrado o no está suspendido.")

    
            elif opcion == "4":
                try:
                    with sqlite3.connect("Evidencia3.db") as conn:
                        mi_cursor = conn.cursor()
                        while True:

                            print("Consultas y reportes de clientes:")
                            print("1. Búsqueda por clave del Cliente")
                            print("2. Búsqueda por nombre del Cliente")
                            print("3. Listado de clientes registrados")
                            print("4. Volver al menú anterior")
                            sub_opcion = input("Ingrese la opción deseada: ")

                            if sub_opcion == "1":

                                clave_buscar = input("Ingrese la clave del cliente a consultar: ")
                                mi_cursor.execute("SELECT ClaveC, NombreCompleto, RFC, CORREO FROM Clientes WHERE ClaveC = ?", (clave_buscar,))
                                cliente = mi_cursor.fetchone()
                                
                                if cliente:
                                    ClaveC, NombreCompleto, RFC, CORREO = cliente
                                    print(f"Detalle del cliente (Clave: {ClaveC}, Nombre Completo: {NombreCompleto}):")
                                    print(f"RFC: {RFC}")
                                    print(f"Correo: {CORREO}")
                                else:
                                    print("La clave de cliente no se encuentra en el sistema.")

                            elif sub_opcion == "2":

                                nombre_buscar = input("Ingrese el nombre del cliente a buscar: ")
                                mi_cursor.execute("SELECT ClaveC, NombreCompleto, RFC, CORREO FROM Clientes WHERE UPPER(NombreCompleto) = UPPER(?)", (nombre_buscar,))
                                clientes = mi_cursor.fetchall()
                                
                                if not clientes:
                                    print("No se encontraron clientes con ese nombre.")
                                else:
                                    print("Búsqueda por nombre de cliente:")
                                    for ClaveC, NombreCompleto, RFC, CORREO in clientes:
                                        print(f"Clave: {ClaveC}, Nombre Completo: {NombreCompleto}, RFC: {RFC}, Correo: {CORREO}")

                            elif sub_opcion == "3":

                                print("Opciones de ordenamiento:")
                                print("1. Ordenado por clave")
                                print("2. Ordenado por nombre")
                                print("3. Volver al Menu Clientes")
                                ordenamiento = input("Seleccione una opción de ordenamiento: ")
                                try:

                                    mi_cursor = conn.cursor()
                                    
                                    if ordenamiento == "1":
                                        mi_cursor.execute("SELECT ClaveC, NombreCompleto FROM Clientes ORDER BY ClaveC")
                                        clientes = mi_cursor.fetchall()
                                        
                                        if not clientes:
                                            print("No hay clientes registrados.")
                                        else:
                                            print("Listado de clientes registrados ordenado por clave:")
                                            for ClaveC, NombreCompleto in clientes:
                                                print(f"Clave: {ClaveC}, Nombre: {NombreCompleto}")
                                            
                                            Resultado = input("¿Desea exportar este resultado a CSV o Excel? (SI/NO): ")
                                            if Resultado.upper() == "SI":
                                                exportar = input("En qué desearías exportarlo, en [A]CSV o [B]Excel: ")
                                                if exportar.upper() == "A":

                                                    fecha_actual = datetime.datetime.now().strftime("%d_%m_%Y")
                                                    nombre_archivo = (f"ReporteClientesActivosPorClave_{fecha_actual}.csv")
                                                    with open(nombre_archivo, mode='w', newline='') as archivo_csv:
                                                        writer = csv.writer(archivo_csv)
                                                        writer.writerow(["Clave", "Nombre"])  
                                                        for ClaveC, NombreCompleto in clientes:
                                                            writer.writerow([ClaveC, NombreCompleto])
                                                    print(f"Resultado exportado a {nombre_archivo}")
                                                elif exportar.upper() == "B":

                                                    fecha_actual = datetime.datetime.now().strftime("%d_%m_%Y")
                                                    nombre_archivo = f"ReporteClientesActivosPorClave_{fecha_actual}.xlsx"
                                                    df = pd.DataFrame(clientes, columns=["Clave", "Nombre"])
                                                    df.to_excel(nombre_archivo, index=False)
                                                    print(f"Resultado exportado a {nombre_archivo}")
                                                else:
                                                    print("Introduce una letra que se encuentra al lado de las opciones")
                                            elif Resultado.upper() == "NO":
                                                break
                                            else:
                                                print("Seleccione 'SI' o 'NO' dependiendo de lo que usted quiere")

                                    elif ordenamiento == "2":

                                        mi_cursor.execute("SELECT NombreCompleto, ClaveC FROM Clientes ORDER BY NombreCompleto")
                                        clientes = mi_cursor.fetchall()
                                        
                                        if not clientes:
                                            print("No hay clientes registrados.")
                                        else:
                                            print("Listado de clientes registrados ordenado por nombre:")
                                            for NombreCompleto, ClaveC in clientes:
                                                print(f"Nombre: {NombreCompleto}, Clave: {ClaveC}")
                                            
                                            Resultado = input("¿Desea exportar este resultado a CSV o Excel? (SI/NO): ")
                                            if Resultado.upper() == "SI":
                                                exportar = input("En qué desearías exportarlo, en [A]CSV o [B]Excel: ")
                                                if exportar.upper() == "A":
                                                    fecha_actual = datetime.datetime.now().strftime("%d_%m_%Y")
                                                    nombre_archivo = (f"ReporteClientesActivosPorNombre_{fecha_actual}.csv")
                                                    with open(nombre_archivo, mode='w', newline='') as archivo_csv:
                                                        writer = csv.writer(archivo_csv)
                                                        writer.writerow(["Nombre", "Clave"])
                                                        for NombreCompleto, ClaveC in clientes:
                                                            writer.writerow([NombreCompleto, ClaveC])
                                                    print(f"Resultado exportado a {nombre_archivo}")
                                                elif exportar.upper() == "B":
                                                    fecha_actual = datetime.datetime.now().strftime("%d_%m_%Y")
                                                    nombre_archivo = f"ReporteClientesActivosPorNombre_{fecha_actual}.xlsx"
                                                    df = pd.DataFrame(clientes, columns=["Nombre", "Clave"])
                                                    df.to_excel(nombre_archivo, index=False)
                                                    print(f"Resultado exportado a {nombre_archivo}")
                                                else:
                                                    print("Introduce una letra que se encuentra al lado de las opciones")
                                            elif Resultado.upper() == "NO":
                                                break
                                            else:
                                                print("Seleccione 'SI' o 'NO' dependiendo de lo que usted quiere")
                                    elif ordenamiento == "3":
                                        break
                                    else:
                                        print("Seleccione una Opcion marcada por el Numero porfavor")
                                except Error as e:
                                    print(e)
                                except Exception:
                                    print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")

                            elif sub_opcion == "4":
                                break
                            else:
                                print("Opción no válida. Por favor, seleccione una opción válida.")
                finally:
                    conn.close
            
            elif opcion == "3":
                break
            else:
                print("Opción no válida. Por favor, seleccione una opción válida.")

    elif MenuPrincipal == "3":
        try:
            with sqlite3.connect("Evidencia3.db") as conn:
                mi_cursor = conn.cursor()
                while True:
                    print("Menú Servicios:")
                    print("1. Agregar un Servicio")
                    print("2. Cancelar un Servicio")
                    print("3. Recuperar un Servicio")
                    print("4. Consultas y reportes de Servicios")
                    print("3. Volver al menú principal")
                    
                    opcion = input("Ingrese la opción deseada: ")

                    if opcion == "1":
                        nombre_servicio = input("Ingrese el nombre del servicio: ")
                        costo_servicio = input("Ingrese el costo del servicio: ")
                        
                        if not nombre_servicio.strip():
                            print("El nombre del servicio no puede quedar vacío.")
                            continue
                        
                        try:
                            costo_servicio = float(costo_servicio)
                            if costo_servicio <= 0.0:
                                print("El costo del servicio debe ser superior a 0.00.")
                                continue
                        except ValueError:
                            print("El costo del servicio debe ser un número válido.")
                            continue

                        mi_cursor.execute("INSERT INTO Servicios (Nombre, Costo) VALUES (?, ?)",
                                    (nombre_servicio, costo_servicio))
                        conn.commit()
                        print("El servicio ha sido registrado con éxito.")

                    elif opcion == "2":
                      mi_cursor.execute("SELECT ClaveS, Nombre FROM Servicios WHERE Estado_Cliente = 0", (cliente_cancelar,))
                      Servicio = mi_cursor.fetchall()
                      if not Servicio:
                        print("No hay Servicios registrados en este momento.")
                      else:
                        print("Servicios Suspendido: ")
                        for ClaveS, Nombre in servicio:
                          print(f"Clave: {ClaveC}, Nombre: {NombreCompleto}")
                        servicio_cancelar = input("Ingrese la clave del Servicio desea Suspender: ")
                        
                        if servicio_cancelar != "0":
                          mi_cursor.execute("SELECT * FROM Clientes WHERE ClaveC = ? AND Estado_Cliente = 1", (servicio_cancelar,))
                          Servicio = mi_cursor.fetchone()
                          if Servicio:
                            print("Datos del Servicio a suspender: ")
                            print(f"Clave: {cliente_suspender[0]}")
                            print(f"Nombre Completo: {cliente_suspender[1]}")
                            print(f"Costo: {cliente_suspender[3]}")
                          
                          
                            confirmar_cancelacion = input("¿Desea cancelar esta nota? (SI/NO): ")
                            
                            if confirmar_cancelacion.upper() == "SI":
                                mi_cursor.execute("UPDATE Servicios SET Estado_Servicio = 1 WHERE ClaveC = ?", (servicio_cancelar,))
                                conn.commit()
                                print("La nota ha sido cancelada exitosamente.")
                            else:
                                print("La nota no ha sido cancelada.")
                        else:
                            print("La nota no existe o ya ha sido cancelada.")
                      

                    elif opcion == "3":
                      mi_cursor.execute("SELECT ClaveS, Nombre FROM Servicios WHERE Estado_Servicio = 0")
                      servicio_recuperar = mi_cursor.fetchall()
              
                      if not servicio_recuperar:
                          print("No hay Servicios suspendidos en este momento.")
                      else:
                          print("Clientes suspendidos:")
                          for ClaveS, Nombre in servicio_recuperar:
                              print(f"Clave: {ClaveS}, Nombre: {Nombre}")
              
                          clave_recuperar = input("Ingrese la clave del cliente que desea recuperar (o 0 para volver al menú anterior): ")
              
                          if clave_recuperar != "0":
                              mi_cursor.execute("SELECT * FROM Clientes WHERE ClaveC = ? AND Estado_Cliente = 1", (clave_recuperar,))
                              servicio_recuperar = mi_cursor.fetchone()
              
                              if servicio_recuperar:
                                  print("Datos del cliente a recuperar:")
                                  print(f"Clave: {servicio_recuperar[0]}")
                                  print(f"Nombre: {servicio_recuperar[1]}")
                                  print(f"Costo: {servicio_recuperar[3]}")
              
                                  confirmacion = input("¿Está seguro de que desea recuperar a este cliente? (SI/NO): ")
                                  if confirmacion.upper() == "SI":
                                      mi_cursor.execute("UPDATE Servicios SET Estado_Servicio = 0 WHERE ClaveS = ?", (clave_recuperar,))
                                      print("Cliente recuperado exitosamente. Ahora puede registrar nuevas notas.")
                                  elif confirmacion.upper() == "NO":
                                      print("Operación cancelada.")
                                  else:
                                      print("Opción no válida.")
                              else:
                                  print("Cliente no encontrado o no está suspendido.")

                
                    elif opcion == "4":
                        while True:

                            print("Consultas y reportes de servicios:")
                            print("1. Búsqueda por clave de servicio")
                            print("2. Búsqueda por nombre de servicio")
                            print("3. Listado de servicios")
                            print("4. Volver al menú anterior")
                            sub_opcion = input("Ingrese la opción deseada: ")
                            
                            if sub_opcion == "1":
                                #mi_cursor.execute("SELECT ClaveS, Nombre FROM Servicios") Preguntar a profesor
                                #servicio = mi_cursor.fetchall()
                                #if not servicio:
                                #    print("No hay servicios registrados.")
                                #else:
                                    #print("Búsqueda por clave de servicio:")
                                    #for ClaveS, Nombre in servicio:
                                        #print(f"Clave: {ClaveS}, Nombre: {Nombre}")
                                
                                clave_buscar = input("Ingrese la clave del servicio a consultar: ")
                                mi_cursor.execute("SELECT ClaveS, Nombre, Costo FROM Servicios WHERE ClaveS = ?", (clave_buscar,))
                                servicio = mi_cursor.fetchone()
                                
                                if servicio:
                                    ClaveS, Nombre, Costo = servicio
                                    print(f"Detalle del servicio (Clave: {ClaveS}, Nombre: {Nombre}):")
                                    print(f"Costo: {Costo}")
                                else:
                                    print("La clave de servicio no se encuentra en el sistema.")


                            elif sub_opcion == "2":

                                nombre_buscar = input("Ingrese el nombre del servicio a buscar: ")
                                mi_cursor.execute("SELECT ClaveS, Nombre, Costo FROM Servicios WHERE UPPER(Nombre) = UPPER(?)", (nombre_buscar,))
                                servicios = mi_cursor.fetchall()
                                
                                if not servicios:
                                    print("No se encontraron servicios con ese nombre.")
                                else:
                                    print("Búsqueda por nombre de servicio:")
                                    for ClaveS, Nombre, Costo in servicios:
                                        print(f"Clave: {ClaveS}, Nombre: {Nombre}, Costo: {Costo}")
                            
                            elif sub_opcion == "3":

                                print("Opciones de ordenamiento:")
                                print("1. Ordenado por clave")
                                print("2. Ordenado por nombre de servicio")
                                print("3. Volver al Menu Servicios")
                                ordenamiento = input("Seleccione una opción de ordenamiento: ")
                                
                                if ordenamiento == "1":

                                    mi_cursor.execute("SELECT ClaveS, Nombre, Costo FROM Servicios ORDER BY ClaveS")
                                    servicios = mi_cursor.fetchall()
                                    
                                    if not servicios:
                                        print("No hay servicios registrados.")
                                    else:
                                        print("Listado de servicios registrados ordenado por clave:")
                                        for ClaveS, Nombre, Costo in servicios:
                                            print(f"Clave: {ClaveS}, Nombre: {Nombre}, Costo: {Costo}")
                                        Resultado = input("¿Desea exportar este resultado a CSV o Excel? (SI/NO): ")
                                        if Resultado.upper() == "SI":
                                            exportar = input("En qué deseas exportarlo, en [A] CSV o [B] Excel: ")
                                            if exportar.upper() == "A":

                                                try:
                                                    fecha_actual = datetime.datetime.now().strftime("%d_%m_%Y")
                                                    nombre_archivo = f"ReporteServiciosPorClave_{fecha_actual}.csv"
                                                    with open(nombre_archivo, mode='w', newline='') as archivo_csv:
                                                        writer = csv.writer(archivo_csv)
                                                        writer.writerow(["Clave", "Nombre", "Costo"])
                                                        for ClaveS, Nombre, Costo in servicios:
                                                            writer.writerow([ClaveS, Nombre, Costo])
                                                    print(f"Resultado exportado a {nombre_archivo}")
                                                except Exception as e:
                                                    print(f"Error al exportar a CSV: {str(e)}")
                                            elif exportar.upper() == "B":

                                                try:
                                                    fecha_actual = datetime.datetime.now().strftime("%d_%m_%Y")
                                                    nombre_archivo = f"ReporteServiciosPorClave_{fecha_actual}.xlsx"
                                                    df = pd.DataFrame(servicios, columns=["Clave", "Nombre", "Costo"])
                                                    df.to_excel(nombre_archivo, index=False)
                                                    print(f"Resultado exportado a {nombre_archivo}")
                                                except Exception as e:
                                                    print(f"Error al exportar a Excel: {str(e)}")
                                            else:
                                                print("Introduce 'A' para CSV o 'B' para Excel.")
                                        elif Resultado.upper() == "NO":
                                            break
                                        else:
                                            print("Selecciona 'SI' o 'NO' según lo que desees.")
                                elif ordenamiento == "2":

                                    mi_cursor.execute("SELECT Nombre, ClaveS, Costo FROM Servicios ORDER BY Nombre")
                                    servicios = mi_cursor.fetchall()
                                    
                                    if not servicios:
                                        print("No hay servicios registrados.")
                                    else:
                                        print("Listado de servicios registrados ordenado por nombre:")
                                        for Nombre, ClaveS, Costo in servicios:
                                            print(f"Nombre: {Nombre}, Clave: {ClaveS}, Costo: {Costo}")

                                        Resultado = input("¿Desea exportar este resultado a CSV o Excel? (SI/NO): ")
                                        if Resultado.upper() == "SI":
                                            exportar = input("En qué deseas exportarlo, en [A] CSV o [B] Excel: ")
                                            if exportar.upper() == "A":

                                                fecha_actual = datetime.datetime.now().strftime("%d_%m_%Y")
                                                nombre_archivo = f"ReporteServiciosPorNombre_{fecha_actual}.csv"
                                                with open(nombre_archivo, mode='w', newline='') as archivo_csv:
                                                    writer = csv.writer(archivo_csv)
                                                    writer.writerow(["Nombre", "Clave", "Costo"])
                                                    for Nombre, ClaveS, Costo in servicios:
                                                        writer.writerow([Nombre, ClaveS, Costo])
                                                print(f"Resultado exportado a {nombre_archivo}")
                                            elif exportar.upper() == "B":

                                                fecha_actual = datetime.datetime.now().strftime("%d_%m_%Y")
                                                nombre_archivo = f"ReporteServiciosPorNombre_{fecha_actual}.xlsx"
                                                df = pd.DataFrame(servicios, columns=["Nombre", "Clave", "Costo"])
                                                df.to_excel(nombre_archivo, index=False)
                                                print(f"Resultado exportado a {nombre_archivo}")
                                            else:
                                                print("Introduce 'A' para CSV o 'B' para Excel.")
                                        elif Resultado.upper() == "NO":
                                            break
                                        else:
                                            print("Selecciona 'SI' o 'NO' según lo que desees.")
                                elif ordenamiento == "3":
                                    break
                                else:
                                    print("Seleccione una Opcion marcada por el Numero porfavor")
                            elif sub_opcion == "4":
                                break
                            else:
                                print("Opción no válida. Por favor, seleccione una opción válida.")
                    
                    elif opcion == "5":
                        break
                    else:
                        print("Opción no válida. Por favor, seleccione una opción válida.")

        except Error as e:
            print(e)
        except Exception:
            print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
        finally:
            conn.close

  
    elif MenuPrincipal == "4":
      while True:
        try:
            with sqlite3.connect("Evidencia3.db") as conn:
                mi_cursor = conn.cursor()
                print("\nMenú Estadísticas:")
                print("1. Servicios más prestados")
                print("2. Clientes con más notas")
                print("3. Promedio de los montos de las notas")
                print("4. Volver al menú principal")
        
                opcion = input("Seleccione una opción: ")
        
                if opcion == "1":
                    try:
                        with sqlite3.connect("Evidencia3.db", detect_types=sqlite3.PARSE_DECLTYPES | sqlite3.PARSE_COLNAMES) as conn:                                                                         
                            cantidad_servicios = int(input("Ingrese la cantidad de servicios más prestados a identificar: "))
                            fecha_inicial_servicios = input("Ingrese la fecha inicial en formato dd/mm/aaaa (o presione Enter para utilizar 01/01/2000): ")
                            fecha_final_servicios = input("Ingrese la fecha final del período a reportar (o presione Enter para utilizar la fecha actual): ")
        
                            if not fecha_inicial_servicios:
                                fecha_inicial_servicios = "01/01/2000"
                            if not fecha_final_servicios:
                                fecha_final_servicios = datetime.datetime.now().strftime("%d/%m/%Y")
        
                            fecha_inicial_servicios = datetime.datetime.strptime(fecha_inicial_servicios, '%d/%m/%Y').date()
                            fecha_final_servicios = datetime.datetime.strptime(fecha_final_servicios, '%d/%m/%Y').date()
        
                            if fecha_inicial_servicios > fecha_final_servicios:
                                print("La fecha inicial no puede ser posterior a la fecha final.")
                            else:
                                mi_cursor.execute("""
                                    SELECT Servicios.Nombre, COUNT(DetalleNotas.ClaveS) AS Cantidad
                                    FROM Notas
                                    JOIN DetalleNotas ON Notas.Folio = DetalleNotas.Folio
                                    JOIN Servicios ON DetalleNotas.ClaveS = Servicios.ClaveS
                                    WHERE Notas.Fecha BETWEEN ? AND ?
                                    GROUP BY Servicios.ClaveS
                                    ORDER BY Cantidad DESC
                                    LIMIT ?
                                """, (fecha_inicial_servicios, fecha_final_servicios, cantidad_servicios))
                                servicios_prestados = mi_cursor.fetchall()
        
                                if not servicios_prestados:
                                    print("No hay servicios prestados en el período seleccionado.")
                                else:
                                    print("\nServicios más prestados en el período seleccionado:")
                                    print("{:<30} {:<10}".format("Nombre del Servicio", "Cantidad"))
                                    for servicio in servicios_prestados:
                                        print("{:<30} {:<10}".format(servicio[0], servicio[1]))
        
                                    exportar_reporte = input("¿Desea exportar este resultado a CSV o Excel? (CSV/Excel/NO): ")
                                    if exportar_reporte.upper() == "CSV":
                                        nombre_archivo = "ReporteServiciosMasPrestados_{}_{}.csv".format(fecha_inicial_servicios.strftime("%d_%m_%Y"), fecha_final_servicios.strftime("%d_%m_%Y"))
                                        with open(nombre_archivo, 'w', newline='') as archivo_csv:
                                            campo_nombres = ['Nombre del Servicio', 'Cantidad']
                                            escritor_csv = csv.DictWriter(archivo_csv, fieldnames=campo_nombres)
                                            escritor_csv.writeheader()
                                            for servicio in servicios_prestados:
                                                escritor_csv.writerow({'Nombre del Servicio': servicio[0], 'Cantidad': servicio[1]})
                                        print(f"Los datos se han exportado a '{nombre_archivo}' exitosamente.")
                                    elif exportar_reporte.upper() == "EXCEL":
                                        nombre_archivo = "ReporteServiciosMasPrestados_{}_{}.xlsx".format(fecha_inicial_servicios.strftime("%d_%m_%Y"), fecha_final_servicios.strftime("%d_%m_%Y"))
                                        workbook = Workbook()
                                        sheet = workbook.active
                                        sheet.append(['Nombre del Servicio', 'Cantidad'])
                                        for servicio in servicios_prestados:
                                            sheet.append([servicio[0], servicio[1]])
                                        for col in sheet.columns:
                                            max_length = 0
                                            column = col[0].column_letter
                                            for cell in col:
                                                try:
                                                    if len(str(cell.value)) > max_length:
                                                        max_length = len(cell.value)
                                                except:
                                                    pass
                                            adjusted_width = (max_length + 2)
                                            sheet.column_dimensions[column].width = adjusted_width
        
                                        workbook.save(nombre_archivo)
                                        print(f"Los datos se han exportado a '{nombre_archivo}' exitosamente.")
                                    elif exportar_reporte.upper() == "NO":
                                        pass
                                    else:
                                        print("Opción de exportación no válida.")
                    finally:
                        conn.close()
        
        
                elif opcion == "2":
                    try:
                        with sqlite3.connect("Evidencia3.db", detect_types=sqlite3.PARSE_DECLTYPES | sqlite3.PARSE_COLNAMES) as conn:
                            cantidad_clientes = int(input("Ingrese la cantidad de clientes con más notas a identificar: "))
                            fecha_inicial_clientes = input("Ingrese la fecha inicial en formato dd/mm/aaaa (o presione Enter para utilizar 01/01/2000): ")
                            fecha_final_clientes = input("Ingrese la fecha final del período a reportar (o presione Enter para utilizar la fecha actual): ")
        
                            if not fecha_inicial_clientes:
                                fecha_inicial_clientes = "01/01/2000"
                            if not fecha_final_clientes:
                                fecha_final_clientes = datetime.datetime.now().strftime("%d/%m/%Y")
        
                            fecha_inicial_clientes = datetime.datetime.strptime(fecha_inicial_clientes, '%d/%m/%Y').date()
                            fecha_final_clientes = datetime.datetime.strptime(fecha_final_clientes, '%d/%m/%Y').date()
        
                            if fecha_inicial_clientes > fecha_final_clientes:
                                print("La fecha inicial no puede ser posterior a la fecha final.")
                            else:
                                mi_cursor.execute("""
                                    SELECT Clientes.NombreCompleto, COUNT(Notas.Folio) AS Cantidad
                                    FROM Notas
                                    JOIN Clientes ON Notas.ClaveC = Clientes.ClaveC
                                    WHERE Notas.Fecha BETWEEN ? AND ?
                                    GROUP BY Clientes.ClaveC
                                    ORDER BY Cantidad DESC
                                    LIMIT ?
                                """, (fecha_inicial_clientes, fecha_final_clientes, cantidad_clientes))
                                clientes_con_mas_notas = mi_cursor.fetchall()
        
                                if not clientes_con_mas_notas:
                                    print("No hay clientes con notas en el período seleccionado.")
                                else:
                                    print("\nClientes con más notas en el período seleccionado:")
                                    print("{:<30} {:<10}".format("Nombre del Cliente", "Cantidad de Notas"))
                                    for cliente in clientes_con_mas_notas:
                                        print("{:<30} {:<10}".format(cliente[0], cliente[1]))
        
                                    exportar_reporte_clientes = input("¿Desea exportar este resultado a CSV o Excel? (CSV/Excel/NO): ")
                                    if exportar_reporte_clientes.upper() == "CSV":
                                        nombre_archivo_clientes = "ReporteClientesConMasNotas_{}_{}.csv".format(fecha_inicial_clientes.strftime("%d_%m_%Y"), fecha_final_clientes.strftime("%d_%m_%Y"))
                                        with open(nombre_archivo_clientes, 'w', newline='') as archivo_csv:
                                            campo_nombres_clientes = ['Nombre del Cliente', 'Cantidad de Notas']
                                            escritor_csv_clientes = csv.DictWriter(archivo_csv, fieldnames=campo_nombres_clientes)
                                            escritor_csv_clientes.writeheader()
                                            for cliente in clientes_con_mas_notas:
                                                escritor_csv_clientes.writerow({'Nombre del Cliente': cliente[0], 'Cantidad de Notas': cliente[1]})
                                        print(f"Los datos se han exportado a '{nombre_archivo_clientes}' exitosamente.")
                                    elif exportar_reporte_clientes.upper() == "EXCEL":
                                        nombre_archivo_clientes = "ReporteClientesConMasNotas_{}_{}.xlsx".format(fecha_inicial_clientes.strftime("%d_%m_%Y"), fecha_final_clientes.strftime("%d_%m_%Y"))
                                        workbook_clientes = Workbook()
                                        sheet_clientes = workbook_clientes.active
                                        sheet_clientes.append(['Nombre del Cliente', 'Cantidad de Notas'])
                                        for cliente in clientes_con_mas_notas:
                                            sheet_clientes.append([cliente[0], cliente[1]])
        
                                        for col_clientes in sheet_clientes.columns:
                                            max_length_clientes = 0
                                            column_clientes = col_clientes[0].column_letter
                                            for cell_clientes in col_clientes:
                                                try:
                                                    if len(str(cell_clientes.value)) > max_length_clientes:
                                                        max_length_clientes = len(cell_clientes.value)
                                                except:
                                                    pass
                                            adjusted_width_clientes = (max_length_clientes + 2)
                                            sheet_clientes.column_dimensions[column_clientes].width = adjusted_width_clientes
        
                                        workbook_clientes.save(nombre_archivo_clientes)
                                        print(f"Los datos se han exportado a '{nombre_archivo_clientes}' exitosamente.")
                                    elif exportar_reporte_clientes.upper() == "NO":
                                        pass
                                    else:
                                        print("Opción de exportación no válida.")
                    finally:
                        conn.close()
        
                elif opcion == "3":
                    try:
                        with sqlite3.connect("Evidencia3.db", detect_types=sqlite3.PARSE_DECLTYPES | sqlite3.PARSE_COLNAMES) as conn:
                            fecha_inicial = input("Ingrese la fecha inicial en formato dd/mm/aaaa (o presione Enter para utilizar 01/01/2000): ")
                            fecha_final = input("Ingrese la fecha final en formato dd/mm/aaaa (o presione Enter para utilizar la fecha actual): ")
        
                            try:
                                if not fecha_inicial:
                                    fecha_inicial = "01/01/2000"
                                if not fecha_final:
                                    fecha_final = datetime.datetime.now().strftime("%d/%m/%Y")
        
                                fecha_inicial = datetime.datetime.strptime(fecha_inicial, '%d/%m/%Y').date()
                                fecha_final = datetime.datetime.strptime(fecha_final, '%d/%m/%Y').date()
        
                                if fecha_inicial > fecha_final:
                                    print("La fecha inicial no puede ser posterior a la fecha final.")
                                else:
                                    mi_cursor.execute("""
                                        SELECT AVG(Monto)
                                        FROM DetalleNotas
                                        WHERE Folio IN (
                                            SELECT Folio
                                            FROM Notas
                                            WHERE Fecha BETWEEN ? AND ?
                                        )
                                    """, (fecha_inicial, fecha_final))
                                    promedio_montos = mi_cursor.fetchone()[0]
        
                                    if promedio_montos is not None:
                                        print(f"Promedio de los montos de las notas en el período: {promedio_montos:.2f}")
                                    else:
                                        print("No hay notas emitidas para el período seleccionado.")
        
                            except ValueError:
                                print("Formato de fecha incorrecto. Introduce la fecha en el formato dd/mm/aaaa.")
        
                    finally:
                        conn.close()
        
                elif opcion == "4":
                    continue
                else:
                    print("Opción no válida. Intente de nuevo.")
        
        except Error as e:
            print(e)
        except Exception:
            print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
        finally:
            conn.close
      
  
    elif MenuPrincipal == "5":
        break
    else:
        print("Opción no válida. Seleccione 1, 2, 3 o 4.")
